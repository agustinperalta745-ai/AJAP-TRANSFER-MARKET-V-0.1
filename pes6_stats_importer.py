"""Import the original PES 6 player database into AJAP.

This module is deliberately data-source agnostic: it accepts the historical
PES/WE spreadsheets as CSV, XLSX or XLS and imports ONLY players that already
exist in AJAP's roster_players table. Fine-grained PES 6 attributes are never
inferred from AJAP OVR.

Preferred data locations:
  data/pes6_original_stats.csv
  data/pes6_original_stats.xlsx
  data/pes6_original_stats.xls
or set PES6_STATS_FILE to a file/directory path.
"""

from __future__ import annotations

import csv
import os
import re
import unicodedata
from pathlib import Path


STAT_COLUMNS = (
    "attack",
    "defence",
    "body_balance",
    "stamina",
    "top_speed",
    "acceleration",
    "response",
    "agility",
    "dribble_accuracy",
    "dribble_speed",
    "short_pass_accuracy",
    "short_pass_speed",
    "long_pass_accuracy",
    "long_pass_speed",
    "shot_accuracy",
    "shot_power",
    "shot_technique",
    "free_kick_accuracy",
    "curling",
    "header",
    "jump",
    "technique",
    "aggression",
    "mentality",
    "gk_skills",
    "teamwork",
)


def _key(value) -> str:
    text = unicodedata.normalize("NFKD", str(value or ""))
    text = "".join(ch for ch in text if not unicodedata.combining(ch))
    text = text.casefold().replace("&", " and ")
    return re.sub(r"[^a-z0-9]+", "", text)


HEADER_ALIASES = {
    "name": (
        "name", "player", "playername", "player name", "nome", "spieler",
    ),
    "attack": ("attack", "att", "offence", "offense"),
    "defence": ("defence", "defense", "def", "df"),
    "body_balance": ("body balance", "bodybalance", "balance", "bal"),
    "stamina": ("stamina", "sta"),
    "top_speed": ("top speed", "topspeed", "speed", "ts"),
    "acceleration": ("acceleration", "accel", "acc"),
    "response": ("response", "responsiveness", "res", "rea"),
    "agility": ("agility", "agi"),
    "dribble_accuracy": ("dribble accuracy", "dribbleaccuracy", "drib accuracy", "da"),
    "dribble_speed": ("dribble speed", "dribblespeed", "drib speed", "ds"),
    "short_pass_accuracy": ("short pass accuracy", "shortpassaccuracy", "short pass acc", "spa"),
    "short_pass_speed": ("short pass speed", "shortpassspeed", "sps"),
    "long_pass_accuracy": ("long pass accuracy", "longpassaccuracy", "long pass acc", "lpa"),
    "long_pass_speed": ("long pass speed", "longpassspeed", "lps"),
    "shot_accuracy": ("shot accuracy", "shotaccuracy", "shot acc", "sa"),
    "shot_power": ("shot power", "shotpower", "sp"),
    "shot_technique": ("shot technique", "shottechnique", "shot tech", "st"),
    "free_kick_accuracy": ("free kick accuracy", "freekickaccuracy", "free kick acc", "fka", "fk accuracy"),
    "curling": ("curling", "curl", "swerve"),
    "header": ("header", "header accuracy", "headeraccuracy", "heading"),
    "jump": ("jump", "jumping"),
    "technique": ("technique", "tech", "tec"),
    "aggression": ("aggression", "agg"),
    "mentality": ("mentality", "mental", "men"),
    "gk_skills": ("gk skills", "gkskills", "keeper skills", "keeperskills", "goalkeeper skills", "goalkeeperskills"),
    "teamwork": ("teamwork", "team work", "tw"),
}

_ALIAS_LOOKUP = {
    _key(alias): canonical
    for canonical, aliases in HEADER_ALIASES.items()
    for alias in aliases
}

# Only use aliases when the original PES/WE spelling is known to differ from the
# AJAP display name. Dynamic matching below handles most surname-only entries.
PLAYER_ALIASES = {
    "riquelme": "Juan Román Riquelme",
    "juninho": "Juninho Pernambucano",
}


def _stat_value(value):
    if value is None or isinstance(value, bool):
        return None
    if isinstance(value, (int, float)):
        number = int(value)
    else:
        match = re.search(r"(?<!\d)(\d{1,2})(?!\d)", str(value).strip())
        if not match:
            return None
        number = int(match.group(1))
    return number if 1 <= number <= 99 else None


def _header_map(row):
    mapping = {}
    for index, value in enumerate(row):
        canonical = _ALIAS_LOOKUP.get(_key(value))
        if canonical and canonical not in mapping:
            mapping[canonical] = index
    return mapping


def _find_header(rows):
    for index, row in enumerate(rows[:30]):
        mapping = _header_map(row)
        if "name" in mapping and sum(1 for stat in STAT_COLUMNS if stat in mapping) >= 5:
            return index, mapping
    return None, None


def _read_csv(path: Path):
    raw = path.read_bytes()
    text = None
    for encoding in ("utf-8-sig", "cp1252", "latin-1"):
        try:
            text = raw.decode(encoding)
            break
        except UnicodeDecodeError:
            pass
    if text is None:
        return []
    try:
        dialect = csv.Sniffer().sniff(text[:8192], delimiters=",;\t|")
    except csv.Error:
        dialect = csv.excel
    return list(csv.reader(text.splitlines(), dialect))


def _read_xlsx(path: Path):
    from openpyxl import load_workbook

    workbook = load_workbook(path, read_only=True, data_only=True)
    result = []
    for sheet in workbook.worksheets:
        result.append((sheet.title, [list(row) for row in sheet.iter_rows(values_only=True)]))
    workbook.close()
    return result


def _read_xls(path: Path):
    import xlrd

    workbook = xlrd.open_workbook(path)
    return [
        (sheet.name, [sheet.row_values(i) for i in range(sheet.nrows)])
        for sheet in workbook.sheets()
    ]


def _sources(base_dir: Path):
    configured = os.getenv("PES6_STATS_FILE", "").strip()
    candidates = []
    if configured:
        target = Path(configured)
        if not target.is_absolute():
            target = base_dir / target
        if target.is_dir():
            candidates.extend(
                p for p in target.rglob("*") if p.suffix.casefold() in (".csv", ".xlsx", ".xls")
            )
        elif target.exists():
            candidates.append(target)

    data_dir = base_dir / "data"
    for name in (
        "pes6_original_stats.csv",
        "pes6_original_stats.xlsx",
        "pes6_original_stats.xls",
    ):
        path = data_dir / name
        if path.exists():
            candidates.append(path)

    collection = data_dir / "pes_we_stats_spreadsheets"
    if collection.exists():
        candidates.extend(
            p for p in collection.rglob("*")
            if p.suffix.casefold() in (".csv", ".xlsx", ".xls")
            and "pes6" in _key(p.name)
        )

    unique = []
    seen = set()
    for path in candidates:
        resolved = str(path.resolve())
        if resolved not in seen:
            seen.add(resolved)
            unique.append(path)
    return unique


def _sheet_is_pes6(file_path: Path, sheet_name: str):
    # A file explicitly named PES6 is trusted. In a multi-game workbook, only
    # sheets explicitly labelled PES6 / PES 6 are read.
    file_key = _key(file_path.stem)
    sheet_key = _key(sheet_name)
    if "pes6" in file_key:
        return True
    return "pes6" in sheet_key


def _build_roster_index(runtime):
    with runtime.db() as conn:
        players = conn.execute("SELECT id, name FROM roster_players ORDER BY id").fetchall()

    exact = {}
    aliases = {}
    for player in players:
        normalized = _key(player["name"])
        exact.setdefault(normalized, []).append(player)

        tokens = [
            _key(token)
            for token in re.split(r"[\s/\-]+", str(player["name"]))
            if len(_key(token)) >= 4
        ]
        # Surname-only and distinctive-token matching is allowed only if unique.
        for token in set(tokens):
            aliases.setdefault(token, []).append(player)

    return players, exact, aliases


def _match_player(source_name, exact, aliases):
    normalized = _key(source_name)
    if not normalized:
        return None

    direct = exact.get(normalized, [])
    if len(direct) == 1:
        return direct[0]

    explicit = PLAYER_ALIASES.get(normalized)
    if explicit:
        direct = exact.get(_key(explicit), [])
        if len(direct) == 1:
            return direct[0]

    # Common PES notation: "R. KEANE" / "RIQUELME" / "MALOUDA". Match a
    # distinctive surname/token only when it identifies exactly one AJAP player.
    source_tokens = [
        _key(token)
        for token in re.split(r"[\s./\-]+", str(source_name))
        if len(_key(token)) >= 4
    ]
    for token in reversed(source_tokens):
        possible = aliases.get(token, [])
        if len(possible) == 1:
            return possible[0]
    return None


def _upsert(runtime, player_id, values, source):
    columns = [column for column in STAT_COLUMNS if values.get(column) is not None]
    if not columns:
        return False

    insert_columns = ["player_id", *columns, "source", "updated_at"]
    placeholders = ["?" for _ in insert_columns[:-1]] + ["CURRENT_TIMESTAMP"]
    params = [player_id, *[values[column] for column in columns], source]
    updates = [f"{column} = excluded.{column}" for column in columns]
    updates.extend(["source = excluded.source", "updated_at = CURRENT_TIMESTAMP"])

    sql = (
        f"INSERT INTO pes6_player_attributes ({', '.join(insert_columns)}) "
        f"VALUES ({', '.join(placeholders)}) "
        f"ON CONFLICT(player_id) DO UPDATE SET {', '.join(updates)}"
    )
    with runtime.db() as conn:
        conn.execute(sql, params)
    return True


def _import_rows(runtime, path: Path, sheet_name: str, rows, exact, aliases):
    header_index, mapping = _find_header(rows)
    if mapping is None:
        return 0, 0, []

    matched = 0
    usable_rows = 0
    unmatched = []
    for row in rows[header_index + 1:]:
        name_index = mapping["name"]
        if name_index >= len(row):
            continue
        source_name = str(row[name_index] or "").strip()
        if not source_name:
            continue

        values = {}
        for stat in STAT_COLUMNS:
            column_index = mapping.get(stat)
            if column_index is None or column_index >= len(row):
                continue
            values[stat] = _stat_value(row[column_index])
        if not any(value is not None for value in values.values()):
            continue
        usable_rows += 1

        player = _match_player(source_name, exact, aliases)
        if not player:
            if len(unmatched) < 30:
                unmatched.append(source_name)
            continue

        source = f"PES 6 original • {path.name} • {sheet_name}"
        if _upsert(runtime, player["id"], values, source):
            matched += 1

    return matched, usable_rows, unmatched


def import_pes6_original_stats(runtime):
    base_dir = Path(getattr(runtime, "__file__", __file__)).resolve().parent
    source_files = _sources(base_dir)
    if not source_files:
        print("AJAP PES6 importer: dataset not bundled yet; keeping verified rows already in DB")
        return {"files": 0, "matched": 0, "rows": 0, "unmatched": []}

    _players, exact, aliases = _build_roster_index(runtime)
    total_matched = 0
    total_rows = 0
    unmatched = []
    processed = 0

    for path in source_files:
        suffix = path.suffix.casefold()
        try:
            if suffix == ".csv":
                books = [(path.stem, _read_csv(path))]
            elif suffix == ".xlsx":
                books = _read_xlsx(path)
            elif suffix == ".xls":
                books = _read_xls(path)
            else:
                continue
        except Exception as exc:
            print(f"WARNING AJAP PES6 importer: no se pudo leer {path.name}: {exc}")
            continue

        file_used = False
        for sheet_name, rows in books:
            if not _sheet_is_pes6(path, sheet_name):
                continue
            matched, usable_rows, missing = _import_rows(
                runtime, path, sheet_name, rows, exact, aliases
            )
            if usable_rows:
                file_used = True
            total_matched += matched
            total_rows += usable_rows
            for name in missing:
                if len(unmatched) < 50 and name not in unmatched:
                    unmatched.append(name)
        if file_used:
            processed += 1

    print(
        "AJAP PES6 importer: "
        f"{total_matched} jugador(es) AJAP vinculados desde {processed} archivo(s) "
        f"({total_rows} filas PES6 leídas)"
    )
    if unmatched:
        print("AJAP PES6 importer: ejemplos sin match: " + ", ".join(unmatched[:10]))

    return {
        "files": processed,
        "matched": total_matched,
        "rows": total_rows,
        "unmatched": unmatched,
    }
